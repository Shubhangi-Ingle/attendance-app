from fastapi import FastAPI, Form, File, UploadFile, Request
from fastapi.responses import FileResponse, JSONResponse
import openpyxl
from openpyxl import Workbook
import os
import json
import secrets
from datetime import datetime, timezone, timedelta
import gspread
import cloudinary
import cloudinary.uploader
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload
import io

app = FastAPI()

EXCEL_FILE = "attendance.xlsx"
IST = timezone(timedelta(hours=5, minutes=30))

# ---- Teacher QR session setup ----
TEACHER_PASSWORD = "Cubeage123"   # <-- change this to your own password
SESSION_FILE = "session.json"
# Set this to your live Render URL
BASE_URL = "https://attendance-app-vghw.onrender.com"

def save_session(token, expires_at):
    with open(SESSION_FILE, "w") as f:
        json.dump({
            "token": token,
            "expires_at": expires_at.isoformat()
        }, f)

def load_session():
    if not os.path.exists(SESSION_FILE):
        return None
    with open(SESSION_FILE, "r") as f:
        return json.load(f)

# ---- Google Sheets & Drive setup ----
GOOGLE_SHEET_ID = "1EwFrEzfHCDy1rNNRGBhWoFV4_Gxfw8D8p360eTBvDhU"
CREDENTIALS_FILE = "credentials.json"

# ---- Cloudinary setup ----
CLOUDINARY_CLOUD_NAME = "do2oynvcf"
CLOUDINARY_API_KEY = "124479454635314"
CLOUDINARY_API_SECRET = "3kxzud_8JY1b53A8Jm2tEC1qjHE"

cloudinary.config(
    cloud_name=CLOUDINARY_CLOUD_NAME,
    api_key=CLOUDINARY_API_KEY,
    api_secret=CLOUDINARY_API_SECRET
)

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

def get_google_sheet(subject):
    creds = Credentials.from_service_account_file(
        CREDENTIALS_FILE,
        scopes=SCOPES
    )

    client = gspread.authorize(creds)
    spreadsheet = client.open_by_key(GOOGLE_SHEET_ID)

    try:
        sheet = spreadsheet.worksheet(subject)
    except gspread.WorksheetNotFound:
        sheet = spreadsheet.add_worksheet(
            title=subject,
            rows=1000,
            cols=10
        )

        sheet.append_row([
            "Name",
            "Subject",
            "Date",
            "Time",
            "Photo"
        ])

    return sheet

def upload_photo_to_cloudinary(content: bytes, filename: str) -> str:
    """Uploads photo bytes to Cloudinary, returns the secure URL."""
    try:
        upload_result = cloudinary.uploader.upload(io.BytesIO(content), public_id=os.path.splitext(filename)[0])
        return upload_result["secure_url"]
    except Exception as e:
        print(f"Cloudinary upload failed: {e}")
        return "Upload failed"

# Serve the homepage (index.html)
@app.get("/")
def serve_form():
    return FileResponse("index.html")

# Serve the teacher page
@app.get("/teacher")
def serve_teacher_page():
    return FileResponse("teacher.html")

# Teacher generates a new time-limited QR session (default 24 hours, customizable)
@app.post("/api/generate-session")
async def generate_session(request: Request):
    data = await request.json()
    password = data.get("password", "")
    duration_hours = data.get("duration_hours", 24)

    if password != TEACHER_PASSWORD:
        return JSONResponse({"message": "Incorrect password"}, status_code=401)

    existing = load_session()
    if existing:
        existing_expiry = datetime.fromisoformat(existing["expires_at"])
        if datetime.now(IST) < existing_expiry:
            return JSONResponse({
                "message": f"A QR code is already active until {existing_expiry.strftime('%I:%M %p, %d %b %Y')}. Please wait until it expires before generating a new one."
            }, status_code=400)

    try:
        duration_hours = float(duration_hours)
        if duration_hours <= 0:
            raise ValueError
    except (TypeError, ValueError):
        return JSONResponse({"message": "Invalid duration"}, status_code=400)

    token = secrets.token_urlsafe(12)
    expires_at = datetime.now(IST) + timedelta(hours=duration_hours)
    save_session(token, expires_at)

    attendance_url = f"{BASE_URL}/?session={token}"

    return JSONResponse({
        "attendance_url": attendance_url,
        "expires_at": expires_at.isoformat(),
        "expires_at_readable": expires_at.strftime("%I:%M %p, %d %b %Y")
    })

# Any teacher opening the page can check if a QR is already active right now
@app.get("/api/current-session")
def current_session():
    session = load_session()
    if not session:
        return JSONResponse({"active": False})

    expires_at = datetime.fromisoformat(session["expires_at"])
    if datetime.now(IST) > expires_at:
        return JSONResponse({"active": False})

    attendance_url = f"{BASE_URL}/?session={session['token']}"
    return JSONResponse({
        "active": True,
        "attendance_url": attendance_url,
        "expires_at": session["expires_at"],
        "expires_at_readable": expires_at.strftime("%I:%M %p, %d %b %Y")
    })

# Student's browser checks if their scanned QR session is still valid
@app.get("/api/validate-session")
def validate_session(token: str = ""):
    session = load_session()
    if not session or session["token"] != token:
        return JSONResponse({"valid": False, "message": "Invalid or unrecognized attendance code."})

    expires_at = datetime.fromisoformat(session["expires_at"])
    if datetime.now(IST) > expires_at:
        return JSONResponse({"valid": False, "message": "This attendance code has expired."})

    return JSONResponse({"valid": True})

# Let the teacher download the latest attendance Excel file anytime
@app.get("/download-excel")
def download_excel():
    return FileResponse(
        path=EXCEL_FILE,
        filename="attendance.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.post("/submit")
async def submit_attendance(
    name: str = Form(...),
    subject: str = Form(...),
    session: str = Form(""),
    photo: UploadFile = File(...)
):
    # Validate the QR session before accepting any attendance
    active_session = load_session()
    if not active_session or active_session["token"] != session:
        return JSONResponse({"message": "Invalid attendance code. Please rescan the QR code."}, status_code=400)

    expires_at = datetime.fromisoformat(active_session["expires_at"])
    if datetime.now(IST) > expires_at:
        return JSONResponse({"message": "This attendance code has expired."}, status_code=400)

    # Use IST regardless of the server's own timezone (Render runs in UTC)
    now = datetime.now(IST)
    date_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H:%M:%S")

    # Photo is required (proves the student took a live picture)
    content = await photo.read()
    if not content:
        return JSONResponse({"message": "No photo captured"}, status_code=400)

    # Upload the photo to Cloudinary
    safe_name = name.strip().replace(" ", "_")
    timestamp_str = now.strftime("%Y%m%d_%H%M%S")
    drive_filename = f"{safe_name}_{subject}_{timestamp_str}.jpg"

    try:
        photo_link = upload_photo_to_cloudinary(content, drive_filename)
    except Exception as e:
        print("Drive upload failed:", e)
        photo_link = "Upload failed"

    # Open or create the Excel file
    if os.path.exists(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Subject", "Date", "Time", "Photo"])

    # Add new row (display "View Photo" if upload succeeded)
    display_text = "View Photo" if photo_link != "Upload failed" else photo_link
    ws.append([name, subject, date_str, time_str, display_text])

    # Make the photo link clickable in Excel
    last_row = ws.max_row
    cell = ws.cell(row=last_row, column=5)
    if photo_link != "Upload failed":
        cell.hyperlink = photo_link
        cell.style = "Hyperlink"
        cell.font = openpyxl.styles.Font(color="0000FF", underline="single")

    wb.save(EXCEL_FILE)

    # Also write the same row into the live Google Sheet (its own tab per subject)
    try:
        sheet = get_google_sheet(subject)
        all_values = sheet.get_all_values()
        if not all_values or (len(all_values) == 1 and all_values[0] == ["Name", "Subject", "Date", "Time", "Photo"]):
            sheet.append_row(["Name", "Subject", "Date", "Time", "Photo"])
        display_link = f'=HYPERLINK("{photo_link}", "View Photo")' if photo_link != "Upload failed" else photo_link
        sheet.append_row([name, subject, date_str, time_str, display_link], value_input_option="USER_ENTERED")
    except Exception as e:
        print("Google Sheet write failed:", e)

    return JSONResponse({"message": "Attendance marked successfully"})